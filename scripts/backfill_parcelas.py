"""Preenche source/installment_info nos gastos já importados e cria as parcelas.

Dry-run por padrão. Use --executar para gravar.

  python scripts/backfill_parcelas.py --household 1
  python scripts/backfill_parcelas.py --household 1 --executar
"""
from __future__ import annotations

import argparse
import io
import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))

from app import create_app  # noqa: E402
from app.extensions import db  # noqa: E402
from app.models import DailyExpense, Installment  # noqa: E402

PARCELA_RE = re.compile(r"^(\d+)/(\d+)$")

# A planilha diz 'corrente'; o modelo fala 'conta'.
SOURCE = {"corrente": "conta", "cartao": "cartao"}


def como_data(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def brl(v) -> str:
    v = Decimal(str(v)).quantize(Decimal("0.01"))
    i, _, c = f"{v:,.2f}".partition(".")
    return f"{i.replace(',', '.')},{c}"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--household", type=int, required=True)
    ap.add_argument("--planilha", default="extratos/revisao.xlsx")
    ap.add_argument("--executar", action="store_true")
    args = ap.parse_args()

    rows = list(load_workbook(RAIZ / args.planilha)["revisao"].iter_rows(min_row=2, values_only=True))
    # external_key -> (conta, parcela, descricao, data, valor)
    info = {}
    for data, conta, desc, valor, parcela, destino, *_rest in rows:
        key = _rest[-1]
        info[key] = (conta, str(parcela or "").strip(), str(desc), data, valor, destino)

    app = create_app()
    with app.app_context():
        gastos = list(db.session.scalars(
            db.select(DailyExpense).where(DailyExpense.household_id == args.household)))

        atualizados = sem_info = 0
        por_source = {"conta": 0, "cartao": 0}
        parcelas_marcadas = 0
        for g in gastos:
            dados = info.get(g.external_key)
            if not dados:
                sem_info += 1
                continue
            conta, parcela, *_ = dados
            g.source = SOURCE.get(conta, conta)
            por_source[g.source] = por_source.get(g.source, 0) + 1
            if PARCELA_RE.match(parcela):
                g.installment_info = parcela
                parcelas_marcadas += 1
            atualizados += 1

        print("=" * 62)
        print("BACKFILL DOS GASTOS")
        print("=" * 62)
        print(f"  atualizados ............ {atualizados}")
        print(f"    no cartão ............ {por_source.get('cartao', 0)}")
        print(f"    na conta ............. {por_source.get('conta', 0)}")
        print(f"  marcados como parcela .. {parcelas_marcadas}")
        print(f"  sem correspondência .... {sem_info}")

        # --------------------------------------------------- parcelas (compras)
        compras = {}
        for key, (conta, parcela, desc, data, valor, destino) in info.items():
            m = PARCELA_RE.match(parcela)
            if not m or destino != "diario":
                continue
            d = como_data(data)
            atual, total = int(m.group(1)), int(m.group(2))
            nome = desc.strip()
            anterior = compras.get(nome)
            if anterior is None or d > anterior["data"]:
                compras[nome] = {"data": d, "atual": atual, "total": total,
                                 "valor": abs(Decimal(str(valor)))}

        existentes = {i.name for i in db.session.scalars(
            db.select(Installment).where(Installment.household_id == args.household))}

        novas = []
        for nome, c in sorted(compras.items()):
            if nome in existentes:
                continue
            novas.append(Installment(
                household_id=args.household, name=nome[:120], amount=c["valor"],
                total_installments=c["total"], current_installment=c["atual"],
                reference_month=date(c["data"].year, c["data"].month, 1)))

        print("\n" + "=" * 62)
        print("PARCELAS A CRIAR")
        print("=" * 62)
        ativas = [p for p in novas if p.remaining > 0]
        for p in novas:
            estado = f"faltam {p.remaining}" if p.remaining else "QUITADA"
            print(f"  {p.name[:30]:32} R$ {brl(p.amount):>8}  "
                  f"{p.current_installment}/{p.total_installments}  {estado}")
        comprometido = sum(p.amount for p in ativas)
        devido = sum(p.total_remaining for p in ativas)
        print(f"\n  {len(novas)} compras, {len(ativas)} ainda ativas")
        print(f"  comprometido por mês: R$ {brl(comprometido)}")
        print(f"  dívida restante:      R$ {brl(devido)}")

        if not args.executar:
            db.session.rollback()
            print("\n>> DRY-RUN: nada gravado. Rode com --executar.")
            return

        db.session.add_all(novas)
        db.session.commit()
        print(f"\n>> GRAVADO: {atualizados} gastos atualizados, {len(novas)} parcelas criadas.")


if __name__ == "__main__":
    main()

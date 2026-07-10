"""Extrai as decisões manuais do usuário para um arquivo de overrides.

Compara a planilha gerada com a planilha revisada e salva as diferenças em
extratos/overrides.json, indexadas por external_key. Assim, regerar a planilha
(com regras novas) nunca perde a revisão humana.

Uso: python scripts/extrair_overrides.py extratos/revisao.xlsx extratos/revisao-02.xlsx
"""
from __future__ import annotations

import io
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

RAIZ = Path(__file__).resolve().parent.parent
OVERRIDES = RAIZ / "extratos" / "overrides.json"


def carrega(path: Path) -> dict:
    ws = load_workbook(path)["revisao"]
    return {r[9]: r for r in ws.iter_rows(min_row=2, values_only=True) if r[9]}


def main():
    if len(sys.argv) != 3:
        sys.exit("uso: extrair_overrides.py <gerada.xlsx> <revisada.xlsx>")
    gerada, revisada = carrega(Path(sys.argv[1])), carrega(Path(sys.argv[2]))

    antigos = {}
    if OVERRIDES.exists():
        antigos = json.loads(OVERRIDES.read_text("utf-8"))

    novos = dict(antigos)
    for k, rb in revisada.items():
        ra = gerada.get(k)
        if ra is None:
            continue
        if ra[5] != rb[5] or (ra[6] or "") != (rb[6] or ""):
            novos[k] = {
                "destino": rb[5],
                "categoria": rb[6] or "",
                "_desc": str(rb[2])[:48],
                "_data": str(rb[0]),
                "_valor": float(rb[3]),
            }

    OVERRIDES.write_text(json.dumps(novos, ensure_ascii=False, indent=2), "utf-8")
    print(f"{len(novos)} overrides salvos em {OVERRIDES}")
    print(f"  ({len(novos) - len(antigos)} novos nesta rodada)\n")
    for k, v in sorted(novos.items(), key=lambda x: x[1]["_data"]):
        print(f"  {v['_data']}  {v['_valor']:>9.2f}  {v['_desc'][:40]:42} -> {v['destino']}")


if __name__ == "__main__":
    main()

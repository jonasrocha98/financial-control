"""Importa a planilha revisada para o banco. Dry-run por padrão.

  python scripts/importar_revisao.py --household 1 --user 1              # simula
  python scripts/importar_revisao.py --household 1 --user 1 --executar   # grava

Idempotente: cada lançamento carrega a external_key do extrato, com constraint
única por household. Rodar duas vezes não duplica nada.
"""
from __future__ import annotations

import argparse
import io
import sys
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
RAIZ = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(RAIZ))
sys.path.insert(0, str(RAIZ / "scripts"))

from app import create_app  # noqa: E402
from app.extensions import db  # noqa: E402
from app.models import DailyExpense, FixedExpense, Household, Income, User  # noqa: E402
from ofx_to_planilha import FIXOS  # noqa: E402


def brl(v) -> str:
    v = Decimal(str(v)).quantize(Decimal("0.01"))
    i, _, c = f"{v:,.2f}".partition(".")
    return f"{i.replace(',', '.')},{c}"


def como_data(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", default="extratos/revisao.xlsx")
    ap.add_argument("--household", type=int, required=True)
    ap.add_argument("--user", type=int, required=True)
    ap.add_argument("--desde", default="2026-01", help="AAAA-MM inclusive")
    ap.add_argument("--ate", default="2026-07", help="AAAA-MM inclusive")
    ap.add_argument("--executar", action="store_true", help="grava de verdade")
    args = ap.parse_args()

    app = create_app()
    with app.app_context():
        casa = db.session.get(Household, args.household)
        usuario = db.session.get(User, args.user)
        if not casa or not usuario:
            sys.exit("household ou user inexistente")
        if usuario.household_id != casa.id:
            sys.exit("esse usuário não pertence a essa casa")
        print(f"casa: {casa.name!r}   usuário: {usuario.name!r}")
        print(f"modo: {'GRAVAÇÃO' if args.executar else 'DRY-RUN (nada será gravado)'}\n")

        rows = list(load_workbook(RAIZ / args.planilha)["revisao"].iter_rows(min_row=2, values_only=True))

        # chaves já no banco -> idempotência
        existentes = set(db.session.scalars(
            db.select(DailyExpense.external_key).where(DailyExpense.household_id == casa.id)
        )) | set(db.session.scalars(
            db.select(Income.external_key).where(Income.household_id == casa.id)
        ))
        existentes.discard(None)

        novos_gastos, novos_entradas = [], []
        pulados = defaultdict(int)
        resumo = defaultdict(lambda: defaultdict(Decimal))

        for data, _conta, desc, valor, _parc, destino, categoria, _conf, _mot, key in rows:
            d = como_data(data)
            mes = d.strftime("%Y-%m")
            if not (args.desde <= mes <= args.ate):
                pulados["fora do período"] += 1
                continue
            if destino == "ignorar":
                pulados["ignorar"] += 1
                continue
            if destino == "fixo":
                pulados["fixo (vira FixedExpense)"] += 1
                continue
            if key in existentes:
                pulados["já importado"] += 1
                continue

            v = abs(Decimal(str(valor)))
            desc = str(desc)[:120]
            if destino == "entrada":
                novos_entradas.append(Income(
                    household_id=casa.id, user_id=usuario.id, description=desc,
                    amount=v, date=d, recurring=(categoria == "Salário"),
                    external_key=key))
                resumo[mes]["entradas"] += v
            elif destino == "diario":
                novos_gastos.append(DailyExpense(
                    household_id=casa.id, user_id=usuario.id, description=desc,
                    amount=v, date=d, category=categoria or None, external_key=key))
                resumo[mes]["gastos"] += v
            else:
                pulados[f"destino desconhecido: {destino!r}"] += 1

        # gastos fixos: um por definição canônica, se ainda não existir
        ja_fixos = {f.name for f in db.session.scalars(
            db.select(FixedExpense).where(FixedExpense.household_id == casa.id))}
        novos_fixos = [FixedExpense(household_id=casa.id, name=n, amount=v, active=True)
                       for n, v in FIXOS.items() if n not in ja_fixos]

        # ------------------------------------------------------------- relatório
        print("=" * 66)
        print("SERÁ GRAVADO")
        print("=" * 66)
        print(f"  entradas .............. {len(novos_entradas):>4}")
        print(f"  gastos do dia a dia ... {len(novos_gastos):>4}")
        print(f"  gastos fixos .......... {len(novos_fixos):>4}")
        print()
        print("  gastos fixos:")
        for f in novos_fixos:
            print(f"     {f.name:28} R$ {brl(f.amount):>9}/mês")
        if not novos_fixos:
            print("     (todos já existem)")

        print("\n" + "=" * 66)
        print("POR MÊS")
        print("=" * 66)
        print(f"  {'mês':9} {'entradas':>12} {'gastos':>12}")
        for mes in sorted(resumo):
            print(f"  {mes:9} {brl(resumo[mes]['entradas']):>12} {brl(resumo[mes]['gastos']):>12}")

        cats = defaultdict(Decimal)
        for g in novos_gastos:
            cats[g.category or "(sem categoria)"] += g.amount
        print("\n" + "=" * 66)
        print("GASTOS POR CATEGORIA")
        print("=" * 66)
        for c, v in sorted(cats.items(), key=lambda x: -x[1]):
            print(f"  {c:28} R$ {brl(v):>10}")

        print("\n" + "=" * 66)
        print("PULADOS")
        print("=" * 66)
        for motivo, n in sorted(pulados.items(), key=lambda x: -x[1]):
            print(f"  {n:>4}  {motivo}")

        if not args.executar:
            print("\n>> DRY-RUN: nada foi gravado. Rode de novo com --executar.")
            return

        db.session.add_all(novos_fixos + novos_entradas + novos_gastos)
        db.session.commit()
        print(f"\n>> GRAVADO: {len(novos_entradas)} entradas, {len(novos_gastos)} gastos, "
              f"{len(novos_fixos)} fixos.")


if __name__ == "__main__":
    main()

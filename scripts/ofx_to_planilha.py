"""Lê os OFX de extratos/ e gera uma planilha de revisão (extratos/revisao.xlsx).

Não toca no banco. Você revisa a planilha, corrige o que estiver errado, salva,
e depois roda scripts/importar_revisao.py.

Uso: python scripts/ofx_to_planilha.py
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = Path(__file__).resolve().parent.parent
EXTRATOS = RAIZ / "extratos"
SAIDA = EXTRATOS / "revisao.xlsx"

TRN_RE = re.compile(r"<STMTTRN>(.*?)</STMTTRN>", re.S)
TAG_RE = re.compile(r"<(\w+)>([^<\r\n]*)")
PARCELA_RE = re.compile(r"\s*-?\s*Parcela\s+(\d+)/(\d+)\s*$", re.I)

# ---------------------------------------------------------------------------
# Regras de classificação. A ORDEM IMPORTA: a primeira que casar vence.
# destino: entrada | diario | fixo | ignorar
# ---------------------------------------------------------------------------
CPF_PROPRIO = r"784\.218"  # o CPF do titular, mascarado no OFX

# Gastos fixos canônicos: valor mensal que o app vai cadastrar como FixedExpense.
# Usar um valor constante (e não a soma dos lançamentos do mês) evita que um mês
# em que a cobrança não postou apareça com custo de vida artificialmente menor.
FIXOS = {
    "Aluguel": Decimal("604.00"),
    "Aula de contrabaixo": Decimal("140.00"),
    "Internet": Decimal("90.00"),
    "Streaming TV": Decimal("35.00"),
    "Nubank+": Decimal("29.00"),
    "Telefone": Decimal("10.00"),
}
FIXO_MENSAL = sum(FIXOS.values())  # R$ 908,00

REGRAS: list[tuple[str, str, str, str, bool]] = [
    # (regex, destino, categoria, motivo, confianca_alta)

    # --- Movimentos que NÃO são receita nem despesa -------------------------
    (r"pagamento de fatura|pagamento recebido", "ignorar", "",
     "fatura do cartão (as compras já entram uma a uma)", True),
    (r"resgate rdb", "ignorar", "",
     "saque da caixinha: dinheiro já seu, não é receita", True),
    (r"aplica..o rdb", "ignorar", "",
     "depósito na caixinha: é poupança, não despesa", True),
    (r"cr.dito em conta", "ignorar", "", "cashback Nubank (estorno, não é renda)", True),

    # --- Movimentos com as contas do próprio titular ------------------------
    # A DIREÇÃO define a natureza: o que CHEGA do Bradesco é a renda dele;
    # o que SAI para BIPA/Bradesco é poupança/investimento, não consumo.
    (r"transfer.ncia recebida.*jonas rocha.*bradesco", "entrada", "Salário",
     "renda vinda da conta do Bradesco (salário / férias)", True),
    (rf"transfer.ncia enviada.*jonas rocha.*{CPF_PROPRIO}", "ignorar", "",
     "aporte em conta/investimento seu (BIPA, Bradesco) — não é despesa", True),

    # --- Gastos fixos -------------------------------------------------------
    # Estes lançamentos NÃO viram gasto do dia a dia: são representados por um
    # FixedExpense (valor mensal constante). Importá-los também duplicaria.
    (r"transfer.ncia enviada pelo pix - rodrigo rocha", "fixo", "Aluguel",
     "gasto fixo mensal", True),
    (r"vero s\.a\.", "fixo", "Internet", "conta fixa mensal", True),
    (r"transfer.ncia enviada.*webert", "fixo", "Streaming TV",
     "streaming de TV, R$ 35/mês", True),
    (r"transfer.ncia enviada.*anderson", "fixo", "Aula de contrabaixo",
     "R$ 140/mês, recorrente", True),
    (r"nubank\+", "fixo", "Nubank+", "assinatura fixa", True),
    (r"plano nucel", "fixo", "Telefone", "plano de celular", True),

    # --- Gastos variáveis identificados pelo usuário ------------------------
    (r"transfer.ncia enviada.*michelli", "diario", "Transferência",
     "gasto variável (R$ 20 a R$ 323)", True),
    (r"60108091carlos", "diario", "Cuidados pessoais", "barbeiro", True),

    # --- Contas de consumo --------------------------------------------------
    (r"elektro", "diario", "Energia", "conta de luz — candidata a gasto fixo", True),
    (r"ultragaz", "diario", "Gás", "boleto recorrente", True),

    # --- Categorias de comércio (a ordem importa) ---------------------------
    (r"drogaria|drogal|droga centro|oficial farma", "diario", "Farmácia", "", True),
    (r"fraldas cia|balloonkids", "diario", "Filhos", "", True),
    (r"supermercado|ki barato|varejao|casa de carnes|barreirense|hortiverde|atacadao",
     "diario", "Mercado", "", True),
    (r"posto |auto ban|allpark|nutag|99app", "diario", "Transporte", "", True),
    (r"pizzaria|beercordeiro|burger|hamburgueria|churras|boteco|padaria|srbatata|"
     r"bolos e salgados|doces caseiros|casa da pizza", "diario", "Alimentação fora", "", True),
    (r"recarga de celular|plano nucel", "diario", "Telefone", "", True),
    (r"google|apple\.?com|anthropic|iof de|nubank\+|nu seguro|amazon digital|nuuvem|"
     r"crunchyroll|capcut|blizzard|csgo-skins", "diario", "Assinaturas", "", True),
    (r"mercadolivre|mercado\*|shopee|shein|aliexpress|amazon", "diario", "Compras online", "", True),
    (r"blu \*br oticas|casa da costura|raros confeccoes|multi coisas|papelaria",
     "diario", "Compras", "", True),
    (r"clinica|estudio ", "diario", "Saúde", "confirme a categoria", False),

    # --- Repasses: dinheiro que passa por você mas não é seu -----------------
    (r"transfer.ncia recebida.*clerison", "ignorar", "",
     "dinheiro do Clerison para você repassar — não é sua renda", True),

    # --- Pix para pessoas: precisa da sua revisão ---------------------------
    (r"transfer.ncia enviada pelo pix", "diario", "Transferência",
     "pix enviado — confirme o motivo", False),
    (r"transfer.ncia (recebida|Recebida)", "entrada", "",
     "dinheiro recebido — é renda ou reembolso?", False),
]


def carrega_overrides() -> dict:
    """Decisões manuais do usuário. Sempre vencem as regras automáticas."""
    p = EXTRATOS / "overrides.json"
    if not p.exists():
        return {}
    return json.loads(p.read_text("utf-8"))


@dataclass
class Trn:
    conta: str
    data: str
    valor: Decimal
    memo: str
    fitid: str
    destino: str = ""
    categoria: str = ""
    motivo: str = ""
    confianca: str = "alta"
    parcela: str = ""

    @property
    def descricao(self) -> str:
        """Memo limpo, sem o sufixo de parcela e sem dados bancários pessoais."""
        d = PARCELA_RE.sub("", self.memo)
        d = re.sub(r"\s*-\s*•.*$", "", d)          # remove CPF/agência/conta
        d = re.sub(r"\s*-\s*[A-Z ]{6,}\(\d+\).*$", "", d)
        return d.strip()[:120]

    @property
    def external_key(self) -> str:
        """Hash composto. O FITID sozinho colide entre parcelas e entre contas."""
        base = f"{self.conta}|{self.fitid}|{self.data}|{self.memo}|{self.valor}"
        return hashlib.sha256(base.encode("utf-8")).hexdigest()[:32]


def ler(path: Path) -> str:
    raw = path.read_bytes()
    for enc in ("utf-8", "cp1252", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", "replace")


def parse(path: Path) -> list[Trn]:
    txt = ler(path)
    conta = "cartao" if "<CREDITCARDMSGSRSV1>" in txt else "corrente"
    out = []
    for bloco in TRN_RE.findall(txt):
        c = dict(TAG_RE.findall(bloco))
        dt = c.get("DTPOSTED", "")[:8]
        t = Trn(
            conta=conta,
            data=f"{dt[:4]}-{dt[4:6]}-{dt[6:8]}",
            valor=Decimal(c.get("TRNAMT", "0")),
            memo=c.get("MEMO", "").strip(),
            fitid=c.get("FITID", "").strip(),
        )
        m = PARCELA_RE.search(t.memo)
        if m:
            t.parcela = f"{m.group(1)}/{m.group(2)}"
        out.append(t)
    return out


def classificar(t: Trn) -> None:
    for regex, destino, categoria, motivo, alta in REGRAS:
        if re.search(regex, t.memo, re.I):
            t.destino, t.categoria, t.motivo = destino, categoria, motivo
            t.confianca = "alta" if alta else "BAIXA"
            return
    # Não casou com nenhuma regra
    t.destino = "entrada" if t.valor > 0 else "diario"
    t.categoria = ""
    t.motivo = "não reconhecido — classifique você"
    t.confianca = "BAIXA"


def main():
    arquivos = sorted(EXTRATOS.glob("*.ofx"))
    if not arquivos:
        sys.exit("Nenhum .ofx em extratos/")

    trns: list[Trn] = []
    for f in arquivos:
        trns.extend(parse(f))
    for t in trns:
        classificar(t)

    # A revisão humana sobrepõe as regras
    overrides = carrega_overrides()
    aplicados = 0
    for t in trns:
        ov = overrides.get(t.external_key)
        if ov:
            t.destino = ov["destino"]
            t.categoria = ov["categoria"]
            t.motivo = "revisado por você"
            t.confianca = "alta"
            aplicados += 1
    if overrides:
        print(f"{aplicados}/{len(overrides)} overrides aplicados")

    trns.sort(key=lambda x: (x.data, x.conta))

    # Detecta colisão de external_key (não deveria haver nenhuma)
    chaves = defaultdict(list)
    for t in trns:
        chaves[t.external_key].append(t)
    dups = [v for v in chaves.values() if len(v) > 1]
    if dups:
        print("!! ATENÇÃO: external_key colidiu — o import perderia dados:")
        for grupo in dups:
            for t in grupo:
                print("   ", t.data, t.valor, t.memo)
        sys.exit(1)
    print(f"OK: {len(trns)} transações, {len(chaves)} chaves únicas (nenhuma colisão)")

    # ---------------------------------------------------------------- planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "revisao"
    cabecalho = ["data", "conta", "descricao", "valor", "parcela",
                 "destino", "categoria", "confianca", "motivo", "external_key"]
    ws.append(cabecalho)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="1F2937")
    for i, _ in enumerate(cabecalho, 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    amarelo = PatternFill("solid", fgColor="FFF3CD")
    cinza = PatternFill("solid", fgColor="EEEEEE")

    for t in trns:
        ws.append([t.data, t.conta, t.descricao, float(t.valor), t.parcela,
                   t.destino, t.categoria, t.confianca, t.motivo, t.external_key])
        linha = ws.max_row
        if t.confianca == "BAIXA":
            for col in range(1, len(cabecalho) + 1):
                ws.cell(row=linha, column=col).fill = amarelo
        elif t.destino == "ignorar":
            for col in range(1, len(cabecalho) + 1):
                ws.cell(row=linha, column=col).fill = cinza

    # Dropdown na coluna destino (F)
    dv = DataValidation(type="list", formula1='"entrada,diario,fixo,ignorar"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(f"F2:F{ws.max_row}")

    larguras = {"A": 12, "B": 10, "C": 42, "D": 11, "E": 9,
                "F": 11, "G": 17, "H": 11, "I": 46, "J": 34}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # ---------------------------------------------------------------- resumo
    ws2 = wb.create_sheet("resumo")
    ws2.append(["mês", "entradas", "gastos (diário)", "ignorados", "qtd linhas"])
    for c in ws2[1]:
        c.font = bold

    por_mes = defaultdict(lambda: {"entrada": Decimal(0), "diario": Decimal(0),
                                   "ignorar": Decimal(0), "n": 0})
    for t in trns:
        m = t.data[:7]
        por_mes[m][t.destino if t.destino in ("entrada", "diario", "ignorar") else "diario"] += abs(t.valor)
        por_mes[m]["n"] += 1
    for m, v in sorted(por_mes.items()):
        ws2.append([m, float(v["entrada"]), float(v["diario"]), float(v["ignorar"]), v["n"]])
    for col in "ABCDE":
        ws2.column_dimensions[col].width = 18

    wb.save(SAIDA)
    print(f"planilha gerada: {SAIDA}")
    print()
    baixa = [t for t in trns if t.confianca == "BAIXA"]
    print(f"{len(baixa)} linhas marcadas em AMARELO precisam da sua revisão:")
    for t in baixa:
        print(f"  {t.data} {t.valor:>10}  {t.descricao[:40]:42} -> {t.destino:8} {t.motivo}")

    print("\nResumo por mês (após as regras):")
    for m, v in sorted(por_mes.items()):
        print(f"  {m}: entradas R$ {v['entrada']:>9} | gastos R$ {v['diario']:>9} "
              f"| ignorados R$ {v['ignorar']:>9}")


if __name__ == "__main__":
    main()

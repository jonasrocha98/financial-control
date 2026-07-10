"""Reconstrói overrides.json lendo a coluna 'motivo' escrita à mão pelo usuário.

O motivo é a fonte de verdade: é a explicação humana do que a transação é.
Este script traduz cada motivo em (destino, categoria).

Uso: python scripts/overrides_do_motivo.py extratos/revisao-02.xlsx
"""
from __future__ import annotations

import io
import json
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
RAIZ = Path(__file__).resolve().parent.parent
OVERRIDES = RAIZ / "extratos" / "overrides.json"

# Motivos que EU gerei automaticamente. Se o motivo for um destes, o usuário não
# escreveu nada ali — e a decisão dele está na coluna 'destino'.
DEFAULTS = {
    "pix enviado — confirme o motivo",
    "não reconhecido — classifique você",
    "dinheiro recebido — é renda ou reembolso?",
    "boleto recorrente — que conta é essa?",
    "transferência entre contas suas — confirme",
    "confirme a categoria",
    "boleto recorrente",
}

# (trecho do motivo, destino, categoria).  Primeiro que casar vence.
# 'fixo' = representado por um FixedExpense; não vira gasto do dia a dia.
MAPA = [
    ("salario",              "entrada", "Salário"),
    ("venda",                "entrada", "Venda de eletrônicos"),
    ("formatação de pc",     "entrada", "Serviços de informática"),
    ("dinheiro recebido",    "entrada", "Recebimento"),

    ("investimento",         "ignorar", ""),

    ("streaming",            "fixo",    "Streaming TV"),
    ("straming",             "fixo",    "Streaming TV"),           # typo do usuário
    ("contrabaixo",          "fixo",    "Aula de contrabaixo"),
    ("internet",             "fixo",    "Internet"),

    ("conta de gas",         "diario",  "Gás"),
    ("barbeiro",             "diario",  "Cuidados pessoais"),
    ("cartão da esposa",     "diario",  "Cartão da esposa"),
    ("cartão esposa",        "diario",  "Cartão da esposa"),
    ("pizzaria",             "diario",  "Alimentação fora"),
    ("alimento",             "diario",  "Alimentação"),
    ("roupas",               "diario",  "Vestuário"),
    ("brinquedo",            "diario",  "Filhos"),
    ("fotos",                "diario",  "Lazer"),
    ("lazer",                "diario",  "Lazer"),
    ("parque",               "diario",  "Lazer"),
    ("papelaria",            "diario",  "Compras"),
    ("varejo",               "diario",  "Compras"),
    ("shopping",             "diario",  "Compras"),
    ("gasto carro",          "diario",  "Carro"),
    ("gasto casa",           "diario",  "Casa"),
    ("gasto normal",         "diario",  ""),
]

VAGOS = ("não lembro",)


def main():
    if len(sys.argv) != 2:
        sys.exit("uso: overrides_do_motivo.py <revisada.xlsx>")
    ws = load_workbook(Path(sys.argv[1]))["revisao"]

    ov, vagos, sem_regra = {}, [], []
    for r in ws.iter_rows(min_row=2, values_only=True):
        data, _, desc, valor, _, destino, _, conf, motivo, key = r
        motivo = (motivo or "").strip().lower()
        if conf != "BAIXA":
            continue

        def grava(dest, cat, origem):
            ov[key] = {"destino": dest, "categoria": cat, "_desc": str(desc)[:48],
                       "_data": str(data), "_valor": float(valor),
                       "_motivo": motivo, "_origem": origem}

        escreveu = motivo and motivo not in DEFAULTS

        # 1. O texto que VOCÊ escreveu manda.
        if escreveu and not any(v in motivo for v in VAGOS):
            for trecho, dest, cat in MAPA:
                if trecho in motivo:
                    grava(dest, cat, "motivo")
                    break
            else:
                sem_regra.append((data, valor, desc, motivo))
            continue

        # 2. Você marcou 'não lembro' -> segue amarelo.
        if escreveu:
            vagos.append((data, valor, desc, motivo))
            continue

        # 3. Sem motivo escrito, mas você mudou o destino para 'ignorar'.
        #    Só uma decisão humana põe 'ignorar' aqui — a regra nunca põe.
        if destino == "ignorar":
            grava("ignorar", "", "destino")

    OVERRIDES.write_text(json.dumps(ov, ensure_ascii=False, indent=2), "utf-8")
    print(f"{len(ov)} overrides gravados a partir dos seus motivos\n")

    from collections import defaultdict
    from decimal import Decimal
    g = defaultdict(lambda: [0, Decimal(0)])
    for v in ov.values():
        k = (v["destino"], v["categoria"] or "(sem categoria)")
        g[k][0] += 1
        g[k][1] += abs(Decimal(str(v["_valor"])))
    print(f"{'destino':9} {'categoria':26} {'qtd':>4} {'total':>10}")
    for (d, c), (q, val) in sorted(g.items()):
        print(f"{d:9} {c:26} {q:>4} {val:>10.2f}")

    if vagos:
        print(f"\n{len(vagos)} linhas que VOCÊ marcou como incertas (seguem amarelas):")
        for d, v, desc, m in vagos:
            print(f"  {d}  {v:>9.2f}  {str(desc)[:34]:36} >> {m}")
    if sem_regra:
        print(f"\n{len(sem_regra)} motivos sem regra de tradução:")
        for d, v, desc, m in sem_regra:
            print(f"  {d}  {v:>9.2f}  {str(desc)[:34]:36} >> {m}")


if __name__ == "__main__":
    main()
